import csv
import io
import re

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Image as PdfImage
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

EXPORT_FORMATS = ("csv", "excel", "pdf")

# Tamil (U+0B80-U+0BFF) — used to spot which export cells need Tamil-specific handling
# below (e.g. Item List's Name (Tamil) column): openpyxl's default Calibri has no Tamil
# glyphs at all, and reportlab has neither Tamil glyphs nor any complex-script shaping
# engine, so an unpatched cell would render as tofu boxes, or worse, real-but-wrongly-
# ordered glyphs (see `_tamil_cell_image`'s docstring).
_TAMIL_RANGE = re.compile("[஀-௿]")

# Ships with Windows 8+/Office 2013+ and covers every major Indic script including
# Tamil (unlike Calibri, the workbook's default) — set explicitly on any cell containing
# Tamil text so a reader without the app's own font-substitution logic still sees real
# glyphs rather than tofu boxes.
_EXCEL_TAMIL_FONT = "Nirmala UI"

_MEDIA_TYPES = {
    "csv": "text/csv",
    "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
}
_EXTENSIONS = {"csv": "csv", "excel": "xlsx", "pdf": "pdf"}


def _to_csv(headers: list[str], grid: list[list[object]]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    writer.writerows(grid)
    # utf-8-sig so Excel opens Tamil (and other non-ASCII) names correctly rather than
    # mangling them via a default codepage guess.
    return buffer.getvalue().encode("utf-8-sig")


def _to_excel(title: str, headers: list[str], grid: list[list[object]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] or "Report"  # Excel sheet-name length limit

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in grid:
        ws.append(row)
        row_cells = ws[ws.max_row]
        for cell in row_cells:
            if cell.value is not None and _TAMIL_RANGE.search(str(cell.value)):
                cell.font = Font(name=_EXCEL_TAMIL_FONT)

    for column_cells in ws.columns:
        length = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=10)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 10), 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _tamil_cell_image(text: str, max_height_pt: float = 9.0, max_width_pt: float = 140.0) -> PdfImage:
    """Embeds Tamil text as a small raster image instead of letting reportlab draw it as
    vector text — confirmed necessary, not just a precaution: reportlab has no complex-
    script shaping engine at all (no GSUB/GPOS ligatures, no pre-base-matra reordering),
    so even with a Tamil-capable font registered, its own text layout renders Tamil
    glyphs in the wrong visual order/spacing. Reuses `printing/tamil_raster.py`'s RAQM-
    shaped renderer — the same code path already proven correct for thermal KOT/bill
    printing — rather than re-solving Tamil shaping a second way here.

    Rendered at a fixed high pixel size for crisp downscaling, then fit to
    `max_height_pt`; if that would still overflow `max_width_pt` (a long Tamil name),
    it's shrunk further by width instead so a single cell can never blow out the table's
    column layout.
    """
    from app.printing.tamil_raster import render_tamil_text_image

    img = render_tamil_text_image(text, font_size=40)
    scale = max_height_pt / img.height
    if img.width * scale > max_width_pt:
        scale = max_width_pt / img.width
    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    buffer.seek(0)
    return PdfImage(buffer, width=img.width * scale, height=img.height * scale)


def _to_pdf(title: str, headers: list[str], grid: list[list[object]]) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4) if len(headers) > 5 else A4,
        title=title,
    )
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles["Title"]), Spacer(1, 12)]

    def cell_value(v: object) -> str | PdfImage:
        text = "" if v is None else str(v)
        return _tamil_cell_image(text) if _TAMIL_RANGE.search(text) else text

    table_data: list[list[object]] = [list(headers)] + [[cell_value(v) for v in row] for row in grid]
    style_commands = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f6f4e")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f2f2")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle(style_commands))
    elements.append(table)
    doc.build(elements)
    return buffer.getvalue()


def export_grid(title: str, headers: list[str], grid: list[list[object]], fmt: str) -> tuple[bytes, str, str]:
    """Renders a pre-built header/row grid into csv/excel/pdf bytes — shared by all
    report export endpoints. Headers and rows are supplied by the caller (report_print_
    service.build_report_body + report_body_to_grid) rather than reflected from a
    Pydantic model's raw field names, so every export format shows the exact same
    renamed/no-id column set the printer version already uses (production feedback
    round 4: "keep the same format of the columns... for csv/pdf/excel").
    """
    if fmt not in EXPORT_FORMATS:
        raise ValueError(f"Unsupported export format: {fmt}")

    filename_stem = title.lower().replace(" ", "_")

    if fmt == "csv":
        content = _to_csv(headers, grid)
    elif fmt == "excel":
        content = _to_excel(title, headers, grid)
    else:
        content = _to_pdf(title, headers, grid)

    return content, _MEDIA_TYPES[fmt], f"{filename_stem}.{_EXTENSIONS[fmt]}"
